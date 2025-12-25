import os
import sqlite3
import threading
from functools import wraps
from io import BytesIO

from flask import Flask, render_template, request, session, redirect, url_for, flash, jsonify, send_file
import main
from storage.repository import MovieRepository
from analysis.clustering import ClusteringService
from analysis.graph import GraphService # 关系图谱展示
from utils.logger import logger

# 基础信息
DB_PATH = os.path.join("data", "movie.db")
app = Flask(__name__)
app.secret_key = "douban_secret_key123" # Flask的session密钥

# 初始化数据仓库
# 初始化业务服务
from analysis.vector_service import VectorService
from analysis.llm_service import LLMService

repo = MovieRepository(DB_PATH)
vector_service = VectorService(repo)
llm_service = LLMService(
    api_key="d38eca80-b3ff-4217-8827-18bc7451b042",
    base_url="https://ark.cn-beijing.volces.com/api/v3",
    model="deepseek-v3-250324"
)

# 启动后台线程预加载向量索引
import threading
def preload_vectors():
    try:
        vector_service.build_index()
    except Exception as e:
        print(f"Error preloading vectors: {e}")

threading.Thread(target=preload_vectors, daemon=True).start()


# 全局变量注入
@app.context_processor
def inject_sidebar_tags():
    """注入侧边栏热门标签 (实时从数据库获取，最多9个)"""
    try:
        top_genres = repo.get_top_genres(limit=9) # 获取热门标签，最多9个
    except Exception:
        top_genres = [] # 防止数据库未初始化时报错

    # 预定义颜色，Bootstrap的颜色关键词
    colors = ['secondary']
    
    tags_data = []
    for i, genre in enumerate(top_genres):
        color = colors[i % len(colors)]
        tags_data.append((genre, color))
    
    return dict(sidebar_tags=tags_data)

# 认证装饰器
def login_required(f):
    @wraps(f) # 保留函数元数据
    def decorated_function(*args, **kwargs):
        if not session.get("logged_in"): # 检查是否登录
            return redirect(url_for("login", next=request.url)) # 未登录则重定向到登录页面
        return f(*args, **kwargs)
    return decorated_function


# Web 页面路由
@app.route("/")
def index():
    stats = repo.get_stats() # 获取统计信息
    return render_template("index.html", stats=stats) # 把数据传给模板


# 重定向到index
@app.route("/index")
def home():
    return index()


# 显示电影列表
@app.route("/movie")
def movie():
    page = int(request.args.get("page", 1)) # 获取当前页码
    limit = 50 # 每页显示数量
    movies, total_pages = repo.get_paginated_movies(page, limit)
    return render_template("movie.html", movies=movies, page=page, total_pages=total_pages)


# 显示电影详情
@app.route("/movie/<int:movie_id>")
def movie_detail(movie_id):
    # 1. 获取电影基本信息   
    movie = repo.get_movie_by_id(movie_id)
    if not movie:
        return "未找到电影", 404
        
    # 2. 获取相似推荐 (对比模式)
    # A. 基于内容 (TF-IDF + 关键词)
    cluster_service = ClusteringService(repo)
    rec_tfidf = cluster_service.get_similar_movies(movie_id, n_top=6)

    # B. 基于语义 (Embedding)
    rec_embedding = vector_service.search_by_id(movie_id, top_k=6)
    
    return render_template("detail.html", movie=movie, rec_tfidf=rec_tfidf, rec_embedding=rec_embedding)


# 词云
@app.route("/word")
def word():
    return render_template("cloud.html")


# 生成词云图片
@app.route("/word/generate")
def word_generate():
    """生成动态词云图片。"""
    try:
        import jieba
        import jieba.analyse
        from wordcloud import WordCloud
        import numpy as np
        from PIL import Image

        # 0. 获取词云类型
        wc_type = request.args.get("type", "category") 
        
        # 1. 加载蒙版图
        img_name = 'tree.jpg' if wc_type == 'category' else 'image.jpg' 
        img_path = os.path.join(app.root_path, 'static', 'assets', 'img', img_name)
        img_array = None
        if os.path.exists(img_path):
            img_array = np.array(Image.open(img_path))
            # 修复 JPG 压缩噪点：将接近白色的像素 (>240) 强制转为纯白 (255)
            img_array[img_array > 240] = 255

        # 2. 确定字体路径
        font_candidates = [
            "/mnt/c/Windows/Fonts/msyh.ttc",   # WSL access to Windows Fonts
            "/mnt/c/Windows/Fonts/simhei.ttf", 
            "msyh.ttc", "simhei.ttf",
            "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        ]
        font_path = None
        for f in font_candidates:
            if os.path.exists(f) or (not f.startswith("/") and os.path.exists(os.path.join(app.root_path, f))): 
                 font_path = f
                 break
        
        wc = WordCloud(
            background_color='white',
            mask=img_array,
            font_path=font_path
        )
        
        # 3. 处理文本与生成逻辑
        if wc_type == "intro":
            # 智能提取关键词 (TF-IDF): 专治简介废话
            # topK=300: 提取前300个关键词
            # allowPOS: 只保留名词(n, nz)和动词(v, vn)
            text = repo.get_all_intro_text()
            tags = jieba.analyse.extract_tags(text, topK=300, withWeight=True, allowPOS=('n', 'nz', 'v', 'vn'))
            # 字典 {word: weight} 传入
            wc.generate_from_frequencies(dict(tags))
        else:
            # 默认分类模式: 直接统计频率
            text = repo.get_all_category_text()
            cut = jieba.cut(text)
            string = ' '.join(cut)
            wc.generate_from_text(string)

        # 4. 输出图片
        image = wc.to_image()
        out = BytesIO()
        image.save(out, format='PNG')
        out.seek(0)
        return send_file(out, mimetype='image/png')

    except Exception as e:
        print(f"WordCloud Error: {e}")
        # 返回占位符或错误图片?
        return f"Error generating wordcloud: {e}", 500


# 关于项目
@app.route("/aboutMe")
def aboutMe():
    return render_template("aboutMe.html")


# 帮助
@app.route("/help")
def help():
    return render_template("help.html")


# 搜索
@app.route("/search")
def search():
    keyword = request.args.get("q", "")
    datalist = repo.search_movies(keyword)
    return render_template("search.html", movies=datalist, keyword=keyword)


# 数据分析
@app.route("/analysis")
def analysis():
    # 1. 概览数据
    genre_list = repo.get_genre_statistics()
    country_labels, country_counts = repo.get_country_statistics()
    
    # 2. 评分与年份数据 (合并自 /score)
    score_labels, score_counts = repo.get_score_distribution()
    year_labels, year_counts = repo.get_year_distribution()
    
    return render_template(
        "analysis.html", 
        genre_data=genre_list, 
        country_labels=country_labels, 
        country_counts=country_counts,
        score_labels=score_labels,
        score_counts=score_counts,
        year_labels=year_labels,
        year_counts=year_counts
    )


# 输入/score时重定向到分析页面
@app.route("/score")
def score():
    return redirect(url_for('analysis') + '#trends')

# 输入/cluster时重定向到分析页面
@app.route("/cluster")
def cluster_page():
    return redirect(url_for('analysis') + '#ai-hub')

# 导出数据
@app.route("/export")
def export_data():
    import openpyxl

    movies = repo.get_all_movies()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "豆瓣电影数据"
    

    
    # 表头
    headers = ["ID", "链接", "封面", "片名", "评分", "评价人数", "简介", "年份", "国家/地区", "类型", "导演", "主演"]
    ws.append(headers)
    
    for movie in movies:
        ws.append(movie)
        
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(out, as_attachment=True, download_name="douban_movies.xlsx")


# 导出统计报表
@app.route("/export/stats")
def export_stats():
    import openpyxl
    wb = openpyxl.Workbook()
    # 删除默认 Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    # 1. 评分分布
    ws1 = wb.create_sheet("评分分布")
    ws1.append(["评分", "数量"])
    s_labels, s_counts = repo.get_score_distribution()
    for l, c in zip(s_labels, s_counts):
        ws1.append([float(l), c])
        
    # 2. 年份分布
    ws2 = wb.create_sheet("年份分布")
    ws2.append(["年份", "数量"])
    y_labels, y_counts = repo.get_year_distribution()
    for l, c in zip(y_labels, y_counts):
        ws2.append([l, c])
        
    # 3. 国家/地区分布
    ws3 = wb.create_sheet("国家地区分布")
    ws3.append(["国家/地区", "数量"])
    c_labels, c_counts = repo.get_country_statistics()
    for l, c in zip(c_labels, c_counts):
        ws3.append([l, c])
        
    # 4. 类型分布
    ws4 = wb.create_sheet("类型分布")
    ws4.append(["类型", "数量"])
    genres = repo.get_genre_statistics()
    for g in genres:
        ws4.append([g['name'], g['value']])
        
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="douban_stats_report.xlsx")


# 导出关系图谱数据
@app.route("/export/graph")
def export_graph():
    import openpyxl
    # 复用 GraphService 获取数据
    service = GraphService(repo)
    # 获取全部节点，不限制数量以提供完整数据
    data = service.build_graph(limit_nodes=9999) 
    
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
        
    # Nodes Sheet
    ws_nodes = wb.create_sheet("Nodes (人物)")
    ws_nodes.append(["Id", "Label", "Category (0=Director, 1=Actor)", "SymbolSize (Frequency)"])
    for node in data.get("nodes", []):
        ws_nodes.append([node["id"], node["name"], node["category"], node["symbolSize"]])
        
    # Edges Sheet
    ws_edges = wb.create_sheet("Edges (关系)")
    ws_edges.append(["Source", "Target"])
    for link in data.get("links", []):
        ws_edges.append([link["source"], link["target"]])
        
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="douban_graph_data.xlsx")


# 导出词云关键词
@app.route("/export/keywords")
def export_keywords():
    import jieba.analyse
    import openpyxl
    
    # 提取逻辑与词云生成一致
    text = repo.get_all_intro_text()
    # 提取 Top 500 关键词
    tags = jieba.analyse.extract_tags(text, topK=500, withWeight=True, allowPOS=('n', 'nz', 'v', 'vn'))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "简介关键词 (Top 500)"
    ws.append(["关键词", "权重 (TF-IDF)", "排名"])
    
    for idx, (word, weight) in enumerate(tags, 1):
        ws.append([word, weight, idx])
        
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, as_attachment=True, download_name="douban_keywords.xlsx")


# 认证与管理后台
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        password = request.form.get("password")
        if password == "douban666":
            session["logged_in"] = True
            return redirect(url_for("admin"))
        else:
            flash("密码错误！", "danger")
    return render_template("login.html")


# 退出登录
@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect(url_for("index"))


# 管理后台
@app.route("/admin")
@login_required
def admin():
    status = load_status()
    # 获取当前数据库的所有表名
    try:
        conn = sqlite3.connect(DB_PATH)
        cur = conn.cursor()
        tables = [r[0] for r in cur.execute("SELECT name FROM sqlite_master WHERE type='table' and name != 'sqlite_sequence'").fetchall()]
        conn.close()
    except Exception as e:
        print(f"Error fetching tables: {e}")
        tables = []
    
    return render_template("admin.html", status=status, current_table=repo.table_name, tables=tables)


# 切换数据源
@app.route("/api/switch_table", methods=["POST"])
@login_required
def switch_table():
    new_table = request.form.get("table_name")
    if new_table:
        repo.set_table(new_table)
        flash(f"已切换数据源为: {new_table}", "success")
    return redirect(url_for("admin"))


# 重命名数据源
@app.route("/api/rename_table", methods=["POST"])
@login_required
def rename_table():
    old_name = request.form.get("old_name")
    new_name = request.form.get("new_name")
    
    if not old_name or not new_name:
        flash("表名不能为空", "danger")
        return redirect(url_for("admin"))
        
    try:
        repo.rename_table(old_name, new_name)
        flash(f"成功将表 {old_name} 重命名为 {new_name}", "success")
    except Exception as e:
        flash(f"重命名失败: {str(e)}", "danger")
        
    return redirect(url_for("admin"))

    return redirect(url_for("admin"))


# RAG 智能问答 API
@app.route("/api/rag/search", methods=["POST"])
def api_rag_search():
    data = request.json
    query = data.get("query", "")
    if not query:
        return jsonify([])
        
    try:
        # 1. AI 提纯关键词并解析特定要求 (Query Refinement)
        analysis = llm_service.analyze_query(query)
        keywords = analysis.get("keywords", query)
        requirements = analysis.get("requirements", "")
        filters = analysis.get("filters", {})
        
        # 2. 语义搜索获取电影列表 (使用提纯后的词 + 过滤器)
        movies = vector_service.search(keywords, top_k=5, filters=filters)
        
        # 3. 调用大模型生成最终推荐语 (包含用户要求)
        answer = llm_service.generate_answer(query, movies, requirements)
        
        return jsonify({
            "answer": answer,
            "movies": movies
        })
    except Exception as e:
        logger.error(f"RAG Search failed: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/rag/rebuild", methods=["POST"])
@login_required
def api_rag_rebuild():
    try:
        data = request.json or {}
        target_table = data.get("table_name")
        
        # DEBUG LOGS
        logger.info(f"Received rebuild request. Data: {data}")
        logger.info(f"Target Table: {target_table}")
        logger.info(f"Current CWD: {os.getcwd()}")
        
        # 如果指定了新表，切换过去
        if target_table:
            logger.info(f"Attempting to set table to: {target_table}")
            repo.set_table(target_table)
            logger.info(f"Switched repository to table: {target_table}. New Persistence Config: {repo.table_name}")
        else:
            logger.warning("No table_name provided in request!")

        def task():
            logger.info("Starting background vector index rebuild...")
            try:
                vector_service.build_index(force_refresh=True)
                logger.info("Background vector rebuild finished successfully.")
            except Exception as e:
                logger.error(f"Background vector rebuild failed: {e}")

        # 异步启动重建
        threading.Thread(target=task, daemon=True).start()
        
        return jsonify({"status": "success", "message": f"Index rebuild started for {repo.table_name}. Please check logs for progress."})
    except Exception as e:
        logger.error(f"Rebuild failed: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/tables", methods=["GET"])
@login_required
def api_admin_tables():
    """获取所有数据表列表"""
    try:
        tables = repo.get_all_tables()
        return jsonify({"tables": tables, "current": repo.table_name})
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500
# API 状态与爬虫
STATUS_FILE = "data/status.json"
import json

def save_status(status_data):
    try:
        temp_file = STATUS_FILE + ".tmp"
        with open(temp_file, "w", encoding="utf-8") as f:
            json.dump(status_data, f)
        os.replace(temp_file, STATUS_FILE)
    except Exception as e:
        logger.error(f"Status save failed: {e}")

def load_status():
    if not os.path.exists(STATUS_FILE):
        return {"status": "idle", "current": 0, "total": 0, "message": ""}
    try:
        with open(STATUS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return {"status": "idle", "current": 0, "total": 0, "message": ""}


@app.route("/api/progress")
def api_progress():
    return jsonify(load_status())

@app.route("/api/logs")
@login_required
def api_logs():
    """获取爬虫日志"""
    log_file = os.path.join("logs", "crawler.log")
    if not os.path.exists(log_file):
        return jsonify({"lines": ["暂无日志"]})
    try:
        # 简单读取最后 100 行
        with open(log_file, "r", encoding="utf-8", errors='ignore') as f:
            lines = f.readlines()
            return jsonify({"lines": lines[-100:]})
    except Exception as e:
        return jsonify({"error": str(e)})


# 爬虫
@app.route("/api/crawl", methods=["POST"])
@login_required
def api_crawl():
    if not os.path.exists("data"):
        os.makedirs("data")

    initial_status = {
        "status": "running",
        "current": 0,
        "total": 0,
        "message": "正在初始化..."
    }
    save_status(initial_status)

    data = request.json
    crawl_type = data.get("crawl_type", "top250")
    tag = data.get("tag", "")
    pages = int(data.get("pages", 1))
    pages = int(data.get("pages", 1))
    limit = int(data.get("limit", 200)) # 获取 limit
    sort = data.get("sort", "recommend") # 获取 sort
    sort = data.get("sort", "recommend") # 获取 sort
    # 新增参数
    target_table_arg = data.get("target_table", "").strip()
    append_mode = data.get("append", False)
    
    # 兼容旧逻辑 no_clear
    no_clear = data.get("no_clear", False)
    if no_clear: 
        append_mode = True

    if crawl_type == "tag":
        initial_status["total"] = limit
    else:
        initial_status["total"] = pages
    save_status(initial_status)
    
    # 决定目标表名
    final_table = "movies"
    if target_table_arg:
        final_table = target_table_arg
    elif tag and not append_mode:
        # 旧逻辑: 如果是 tag 且不清空(这里用 append_mode 判断)，则分表
        # 注意: 如果用户没指定 table 且没开启 append，默认也是 movies_tag ? 
        # 为了简化，如果没有指定 table，默认就是 movies
        # 但为了保留原本 "movies_tag" 的特性，我们还是保留一下
        final_table = f"movies_{tag}"
        
    # 自动切换到目标表
    repo.set_table(final_table)
    
    base_url = "https://movie.douban.com/top250"
    if crawl_type == "tag":
        base_url = "JSON_API"
    
    def on_progress(current, total):
        status = load_status()
        status["current"] = current
        status["total"] = total
        if crawl_type == "tag":
            status["message"] = f"正在爬取第 {current}/{total} 部..."
        else:
            status["message"] = f"正在爬取第 {current}/{total} 页..."
        save_status(status)

    def task():
        try:
            logger.info(f"Starting Background Crawl: {crawl_type}, Pages: {pages}, Limit: {limit}, Table: {final_table}, Append: {append_mode}")
            main.run_crawl(
                base_url=base_url,
                tag=tag,
                pages=pages,
                limit=limit, # 传递 limit
                delay=3.0, # 增加延迟防封 (3秒)
                db_path="data/movie.db",
                clear=not append_mode,
                target_table=final_table,
                sort=sort,
                verbose=False,
                progress_callback=on_progress
            )
            logger.info("Background Crawl Finished Successfully.")
            
            status = load_status()
            status["status"] = "finished"
            status["message"] = "爬取完成！数据已更新。"
            if crawl_type == "tag":
                status["current"] = limit
            else:
                status["current"] = pages
            save_status(status)
            
        except Exception as e:
            logger.error(f"Background Crawl Error: {e}")
            status = load_status()
            status["status"] = "error"
            status["message"] = f"发生错误: {str(e)}"
            save_status(status)

    thread = threading.Thread(target=task)
    thread.daemon = True
    thread.start()

    return jsonify({"status": "success", "message": "爬虫已启动..."})



# 聚类路由
@app.route("/api/cluster/data")
def clustering_data():
    try:
        k = int(request.args.get("k", 10)) # 默认为 10 类
        # 限制 k 在 2 到 20 之间，防止错误
        k = max(2, min(k, 20))
        
        service = ClusteringService(repo)
        data = service.perform_clustering(n_clusters=k)
        if data is None:
             return jsonify({"error": "数据不足，无法进行聚类分析 (需要剧情简介数据)"}), 400
        return jsonify(data)
    except Exception as e:
        logger.error(f"Clustering Error: {e}")
        return jsonify({"error": str(e)}), 500


# 知识图谱路由
@app.route("/api/graph/data")

def graph_data():
    try:
        # 默认展示 Top 80 人物，防止图太大卡顿
        limit = int(request.args.get("limit", 80))
        service = GraphService(repo)
        data = service.build_graph(limit_nodes=limit)
        return jsonify(data)
    except Exception as e:
        logger.error(f"Graph Error: {e}")
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(debug=True, port=5002)
